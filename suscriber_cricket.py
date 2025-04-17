import sys
import json
import sqlite3
import os
import threading
import signal
from datetime import datetime
import math  # 🆕 Añadido para manejar NaN

from openpyxl import Workbook, load_workbook

import paho.mqtt.client as mqtt
from PyQt5.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QWidget, QLabel, QComboBox
from PyQt5.QtCore import QTimer
import pyqtgraph as pg
from pyqtgraph import AxisItem

# Config
BROKER = "192.168.0.101"
PORT = 1884
TOPIC = "esp32/sensores/crickethub1"
DB_FILE = "datos_sensores.db"
EXCEL_FILE = "mediciones.xlsx"

# Variables globales para gráfico
x_data = []
y_data = []
selected_variable = "temp"
lock = threading.Lock()

# DB setup
conn = sqlite3.connect(DB_FILE, check_same_thread=False)
cursor = conn.cursor()
cursor.execute("""
CREATE TABLE IF NOT EXISTS mediciones (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    timestamp TEXT,
    accel_x REAL,
    accel_y REAL,
    accel_z REAL,
    lux REAL,
    temp REAL,
    humidity REAL,
    co2 REAL
)
""")
conn.commit()

# Excel setup
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.append(["timestamp", "accel_x", "accel_y", "accel_z", "lux", "temp", "humidity", "co2"])
    wb.save(EXCEL_FILE)

# MQTT callbacks
def on_connect(client, userdata, flags, rc):
    if rc == 0:
        print("✅ Conectado al broker MQTT")
        client.subscribe(TOPIC)
    else:
        print(f"❌ Error de conexión MQTT: {rc}")

def on_message(client, userdata, msg):
    global x_data, y_data
    try:
        payload = json.loads(msg.payload.decode())
        fila = (
            payload["timestamp"],
            float(payload.get("accel_x", 0)),
            float(payload.get("accel_y", 0)),
            float(payload.get("accel_z", 0)),
            float(payload.get("lux", 0)),
            float(payload.get("temp", 0)),
            float(payload.get("humidity", 0)),
            float(payload.get("co2", 0)),
        )

        # Guardar en DB
        cursor.execute("INSERT INTO mediciones (timestamp, accel_x, accel_y, accel_z, lux, temp, humidity, co2) VALUES (?, ?, ?, ?, ?, ?, ?, ?)", fila)
        conn.commit()

        # Guardar en Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Datos"]
        ws.append(fila)
        wb.save(EXCEL_FILE)

        # Actualizar datos para gráfica
        with lock:
            dt = datetime.strptime(payload["timestamp"], "%Y-%m-%dT%H:%M:%S")
            x_data.append(dt)
            y_data.append(payload.get(selected_variable, 0))

            if len(x_data) > 500:
                x_data = x_data[-500:]
                y_data = y_data[-500:]

    except Exception as e:
        print(f"⚠️ Error procesando mensaje: {e}")

# Eje X con etiquetas de tiempo
class TimeAxisItem(AxisItem):
    def tickStrings(self, values, scale, spacing):
        return [datetime.fromtimestamp(value).strftime("%d-%m %H:%M") for value in values]

# GUI
class SensorDashboard(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("📊 Dashboard CricketHub")

        self.graphWidget = pg.PlotWidget(axisItems={'bottom': TimeAxisItem(orientation='bottom')})
        self.graphWidget.setMouseEnabled(x=True, y=True)
        self.graphWidget.showGrid(x=True, y=True)
        self.graphWidget.setBackground('w')
        self.graphWidget.getPlotItem().getAxis("left").setLabel("Valor")
        self.graphWidget.getPlotItem().getAxis("bottom").setLabel("Tiempo")

        self.plot = self.graphWidget.plot([], [], pen=pg.mkPen(color=(0, 100, 200), width=2))

        self.label = QLabel("Variable:")
        self.combo = QComboBox()
        self.combo.addItems(["accel_x", "accel_y", "accel_z", "lux", "temp", "humidity", "co2"])
        self.combo.currentTextChanged.connect(self.change_variable)

        layout = QVBoxLayout()
        layout.addWidget(self.label)
        layout.addWidget(self.combo)
        layout.addWidget(self.graphWidget)

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

        self.timer = QTimer()
        self.timer.timeout.connect(self.update_plot)
        self.timer.start(1000)

    def change_variable(self, value):
        global selected_variable, x_data, y_data
        with lock:
            selected_variable = value
            x_data.clear()
            y_data.clear()

    def update_plot(self):
        try:
            with lock:
                if len(x_data) > 0:
                    x_timestamps = []
                    y_valid = []
                    for t, y in zip(x_data, y_data):
                        try:
                            if not isinstance(y, (int, float)) or math.isnan(y):
                                continue  # 🆕 Ignora valores no numéricos o NaN
                            x_ts = t.timestamp()
                            x_timestamps.append(x_ts)
                            y_valid.append(y)
                        except Exception as e:
                            print(f"⚠️ Error en timestamp: {e}")
                    if x_timestamps and len(x_timestamps) == len(y_valid):
                        self.plot.setData(x=x_timestamps, y=y_valid)
                        self.graphWidget.setTitle(f"Evolución de {selected_variable}")
        except Exception as e:
            print(f"⚠️ Error actualizando gráfica: {e}")

# Threads

def mqtt_thread_func():
    client = mqtt.Client()
    client.on_connect = on_connect
    client.on_message = on_message
    client.connect(BROKER, PORT, 60)
    client.loop_forever()

def gui_thread_func():
    app = QApplication(sys.argv)
    dashboard = SensorDashboard()
    dashboard.show()
    sys.exit(app.exec_())

# Main
if __name__ == "__main__":
    mqtt_thread = threading.Thread(target=mqtt_thread_func)
    gui_thread = threading.Thread(target=gui_thread_func)

    mqtt_thread.start()
    gui_thread.start()

    try:
        mqtt_thread.join()
        gui_thread.join()
    except KeyboardInterrupt:
        print("\n🛑 Finalizando programa...")
        conn.close()
        sys.exit(0)
